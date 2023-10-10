# Import required modules
import subprocess
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from pathlib import Path


class UserCheckLogic():
    def __init__(self, excel_file):
        """
        Initialization of UserCheckLogic class

        Args:
            excel_file (str): Path to Excel file used for user status check.
        """
        # Check if the Excel file exists before opening it
        if not Path(excel_file).is_file():
            raise FileNotFoundError(f"Datoteka '{excel_file}' ne postoji.")
        self.excel_file = excel_file

    def current_time(self):
        """
        Returns current date to be used as part of sheet name.

        Returns:
            datetime.date: Current date.
        """
        current_date = datetime.now().date()
        return current_date

    def check_user_status(self):
        """
        Perform the check of user status based on Excel file and updates it with results.
        """
        # Open the Excel file with openpyxl
        self.workbook = load_workbook(self.excel_file)
        self.sheet = self.workbook.active
        self.sheet.title = f"Provjera_{self.current_time()}"

        # Open it with DataFrame to enable search by column names, regardless of position
        self.df = pd.read_excel(self.excel_file)

        # Check if "Username" column exists
        if "Username" not in self.df.columns:
            raise ValueError("Stupac \"Username\" nije pronađen u Excel datoteci.")

        # Check if "Account Status" column exists
        if "Account Status" not in self.df.columns:
            raise ValueError("Stupac \"Account Status\" nije pronađen u Excel datoteci.")

        # Find the positions of "Username" and "Account Status" columns
        username_col_name = "Username"
        account_status_col_name = "Account Status" 

        # Iterating through rows using DataFrame
        for index, row in self.df.iterrows():
            username_column = row[username_col_name]  # Iterate through "Username" column
            if pd.notna(username_column): # Check if username values are not emtpy
                # Split username to remove domain if present
                username_parts = username_column.split("/")
                username = username_parts[-1]  # Take the last part as username
                
                # Commit net user command
                command = f"net user {username}"  # Add " /domain" or applicable at the end
                result = subprocess.run(command, shell=True, stdout=subprocess.PIPE, text=True)

                # Analyse if the user is active or not or even if the user exist or not       
                if "Account active               No" in result.stdout:
                    account_status = "Disabled"
                elif "Account active               Yes" in result.stdout:
                    account_status = "Active"
                else:
                    account_status = "User not found"

                # Update the Excel file with gathered information
                self.sheet.cell(row=index + 2,
                                column=self.df.columns.get_loc(account_status_col_name) + 1,
                                value=account_status)
            else:
                self.sheet.cell(row=index + 2,
                                column=self.df.columns.get_loc(account_status_col_name) + 1,
                                value="No username provided")
        

    def save_to_file(self):
        """Save to Excel file"""
        self.workbook.save(self.excel_file)